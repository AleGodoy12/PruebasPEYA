import openpyxl
import urllib.request
import json
import time
import random
import ssl
import docx
import openai
import requests
import datetime
from bs4 import BeautifulSoup

# Configurar OpenAI
openai.api_key = "apigpt"
# Desactivar la verificación SSL (solo para pruebas)
ssl._create_default_https_context = ssl._create_unverified_context

MAX_DAILY_REQUESTS = 1000  # Ajusta este número según tu plan de API

def leer_lineamientos(archivo_word):
    doc = docx.Document(archivo_word)
    texto_completo = "\n".join([paragraph.text for paragraph in doc.paragraphs])
    return texto_completo

def aplicar_lineamientos(info_producto, lineamientos):
    prompt = f"""
    Aplica los siguientes lineamientos al producto:
    {lineamientos}
    
    Información del producto:
    {json.dumps(info_producto, ensure_ascii=False, indent=2)}
    
    Devuelve la información del producto actualizada en formato JSON.
    """
    
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "Eres un asistente experto en aplicar lineamientos a información de productos."},
            {"role": "user", "content": prompt}
        ]
    )
    
    try:
        return json.loads(response.choices[0].message.content)
    except json.JSONDecodeError:
        print("Error al decodificar la respuesta de OpenAI. Contenido de la respuesta:")
        print(response.choices[0].message.content)
        return info_producto

def buscar_producto(barcode, api_key, daily_requests):
    max_intentos = 10
    tiempo_espera_base = 5
    
    for intento in range(max_intentos):
        if daily_requests['count'] >= MAX_DAILY_REQUESTS:
            print("Se ha alcanzado el límite diario de solicitudes.")
            return None

        url = f"https://go-upc.com/api/v1/code/{barcode}?key={api_key}"
        print(f"Consultando URL: {url}")
        
        try:
            response = requests.get(url)
            daily_requests['count'] += 1
            response.raise_for_status()
            data = response.json()
            product = data.get('product', {})
            if not product:
                print(f"No se encontró información del producto para el código de barras {barcode}")
                return None
            
            return {
                'title': product.get('name', ''),
                'image': product.get('imageUrl', ''),
                'brand': product.get('brand', ''),
                'content_value': product.get('size', '').split()[0] if product.get('size') else '',
                'content_unit': product.get('size', '').split()[1] if product.get('size') and len(product.get('size').split()) > 1 else '',
                'units_per_pack': '1',
                'additional_image': ''
            }
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 429:
                tiempo_espera = tiempo_espera_base * (2 ** intento) + random.uniform(0, 1)
                print(f"Error 429: Demasiadas solicitudes. Esperando {tiempo_espera:.2f} segundos antes de reintentar...")
                time.sleep(tiempo_espera)
            else:
                print(f"Error HTTP al consultar la API: {str(e)}")
                return None
        except Exception as e:
            print(f"Error al consultar la API: {str(e)}")
            return None
    
    print(f"No se pudo obtener información para el código de barras {barcode} después de {max_intentos} intentos.")
    return None

def crear_resultados(archivo_origen, archivo_destino, api_key, archivo_lineamientos):
    print(f"Leyendo archivo origen: {archivo_origen}")
    wb_origen = openpyxl.load_workbook(archivo_origen)
    sheet_origen = wb_origen.active

    print(f"Creando archivo destino: {archivo_destino}")
    wb_destino = openpyxl.Workbook()
    sheet_destino = wb_destino.active

    # Agregar encabezados
    encabezados = ['id', 'barcode', 'image', 'country', 'field', 'original_value', 'found_value']
    sheet_destino.append(encabezados)

    lineamientos = leer_lineamientos(archivo_lineamientos)

    campos = ['title', 'image', 'brand', 'content_value', 'content_unit', 'units_per_pack', 'additional_image', 'packaging', 'storage_type']

    total_barcodes = 0
    barcodes_encontrados = 0
    barcodes_no_encontrados = 0
    barcodes_procesados = 0

    # Procesar solo los primeros 2 barcodes
    for row in range(2, min(sheet_origen.max_row + 1, 4)):  # 4 porque empezamos en 2 y queremos 2 barcodes
        barcode = sheet_origen.cell(row=row, column=2).value
        if barcode:
            total_barcodes += 1
            try:
                id_producto = sheet_origen.cell(row=row, column=1).value
                imagen = sheet_origen.cell(row=row, column=3).value
                pais = sheet_origen.cell(row=row, column=4).value

                print(f"\nProcesando barcode {barcode}...")
                info_producto = buscar_producto(barcode, api_key)
                
                if info_producto:
                    barcodes_encontrados += 1
                    print("Información encontrada:")
                    print(json.dumps(info_producto, indent=2, ensure_ascii=False))
                    
                    for campo in campos:
                        sheet_destino.append([
                            id_producto,
                            barcode,
                            imagen,
                            pais,
                            campo,
                            info_producto.get(campo, 'x'),
                            info_producto.get(campo, 'x')
                        ])
                else:
                    barcodes_no_encontrados += 1
                    print("No se encontró información para este barcode")
                    for campo in campos:
                        sheet_destino.append([
                            id_producto,
                            barcode,
                            imagen,
                            pais,
                            campo,
                            'x',
                            'x'
                        ])
                
                barcodes_procesados += 1
                print(f"Procesado: {barcodes_procesados}/2 - Barcode: {barcode}")
                
                time.sleep(random.uniform(2, 3))  # Pausa entre solicitudes

            except Exception as e:
                print(f"Error al procesar el barcode {barcode}: {str(e)}")
                barcodes_no_encontrados += 1

    # Guardar resultados
    wb_destino.save(archivo_destino)
    
    print("\nResumen del proceso:")
    print(f"Total de códigos de barras procesados: {total_barcodes}")
    print(f"Códigos de barras encontrados: {barcodes_encontrados}")
    print(f"Códigos de barras no encontrados: {barcodes_no_encontrados}")

# Ejecutar el script
if __name__ == "__main__":
    api_key_goupc = "apigoupc"
    crear_resultados('Test.xlsx', 'Resultados_2_barcodes.xlsx', api_key_goupc)
