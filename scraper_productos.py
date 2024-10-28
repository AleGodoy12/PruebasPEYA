import pandas as pd
import requests
import openpyxl
import time
import random
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import ssl
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.poolmanager import PoolManager
import urllib.request
import json

# Suprimir las advertencias de solicitudes inseguras
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class TLSAdapter(HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(num_pools=connections,
                                       maxsize=maxsize,
                                       block=block,
                                       ssl_version=ssl.PROTOCOL_TLSv1_2)

session = requests.Session()
session.mount("https://", TLSAdapter())

# Desactivar la verificación SSL (solo para pruebas)
ssl._create_default_https_context = ssl._create_unverified_context

def buscar_producto(barcode, api_key):
    url = f"https://go-upc.com/api/v1/code/{barcode}"
    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {api_key}')
    
    for _ in range(3):  # Intentar hasta 3 veces
        try:
            with urllib.request.urlopen(req, timeout=10) as response:
                content = response.read()
                data = json.loads(content.decode())
                product = data.get("product", {})
                return {
                    'Title': product.get('name', ''),
                    'Content': product.get('description', ''),
                    'Brand': product.get('brand', ''),
                    'Category': product.get('category', ''),
                    'Image': product.get('imageUrl', '')
                }
        except urllib.error.URLError as e:
            print(f"Error de conexión: {e}")
            time.sleep(5)  # Esperar 5 segundos antes de reintentar
    
    print(f"No se pudo conectar después de 3 intentos para el barcode {barcode}")
    return None

def actualizar_excel(file_path, api_key):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Iterar sobre las primeras 10 filas del Excel (excluyendo la cabecera)
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
        barcode = row[1]  # Asumiendo que el barcode está en la segunda columna
        
        # Buscar el producto
        info_producto = buscar_producto(barcode, api_key)
        
        if info_producto:
            # Actualizar las celdas correspondientes
            sheet.cell(row=row_index, column=3, value=info_producto['Title'])
            sheet.cell(row=row_index, column=4, value=info_producto['Content'])
            sheet.cell(row=row_index, column=5, value=info_producto['Brand'])
            sheet.cell(row=row_index, column=6, value=info_producto['Category'])
            sheet.cell(row=row_index, column=7, value=info_producto['Image'])
        
        # Esperar un tiempo aleatorio para no sobrecargar el servidor
        time.sleep(random.uniform(1, 3))
    
    # Guardar los cambios
    wb.save(file_path)

# Ejecutar el scraper
api_key = "apigoupc"
actualizar_excel('Test.xlsx', api_key)
