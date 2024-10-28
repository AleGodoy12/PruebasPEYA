import pandas as pd
import openai
from openpyxl import load_workbook

# Configurar la clave de API de OpenAI
openai.api_key = "apigpt"

# Cargar el archivo Excel
wb = load_workbook(filename='Test.xlsx', read_only=True)
sheet = wb.active

# Obtener un resumen del contenido
resumen = f"El archivo Excel 'Test.xlsx' contiene {sheet.max_row} filas y {sheet.max_column} columnas."

# Obtener las primeras filas como muestra
muestra = []
for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
    muestra.append(row)

# Convertir la muestra a un DataFrame de pandas
df_muestra = pd.DataFrame(muestra[1:], columns=muestra[0])

# Crear un prompt para GPT
prompt = f"""
Analiza la siguiente información de un archivo Excel:

{resumen}

Muestra de las primeras filas:
{df_muestra.to_string()}

Basándote en esta información, describe de qué trata este archivo Excel y qué tipo de datos contiene.
"""

# Llamar a la API de OpenAI
response = openai.ChatCompletion.create(
    model="gpt-3.5-turbo",
    messages=[
        {"role": "system", "content": "Eres un asistente experto en análisis de datos."},
        {"role": "user", "content": prompt}
    ]
)

# Imprimir la respuesta
print(response.choices[0].message['content'])

# Cerrar el archivo Excel
wb.close()
